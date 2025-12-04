# logic.py
from openpyxl import load_workbook, Workbook
from pathlib import Path
import shutil
import google.generativeai as genai
from docx2pdf import convert
import os
from copy import copy

api_key = os.getenv("GEMINI_API_KEY")

if not api_key:
    raise ValueError("⚠️ No se encontró la variable de entorno GEMINI_API_KEY. Configúrala antes de ejecutar el programa.")

genai.configure(api_key=api_key)
model = genai.GenerativeModel("gemini-2.5-flash")

def procesar(
    ruta_excel: str,
    ruta_carpeta: str,
    check_box: bool,
    gemi_txt: str,
    columna: int = 4,
    on_log=print,
    stop_event=None  
):
    """
    Pipeline:
    1) Asegura carpeta destino y abre/crea Excel
    2) Copia carpetas listadas en la columna 'columna' del Excel
    3) Inserta columna "Observaciones" y guarda Excel (y copia en carpeta destino)
    4) Sube archivos a Gemini y genera 1 .txt por carpeta/obra en 'respuestas_gemini/'
    5) Devuelve resumen
    """
    try:
        
        def _check_cancel():
            if stop_event is not None and stop_event.is_set():
                on_log("⏹ Proceso cancelado por el usuario.")
                return {
                    "cancelado": True
                }
            return None
        
        rutaExcel = Path(ruta_excel)
        rutaCarpeta = Path(ruta_carpeta)
        col = int(columna)

        on_log(f"📁 Destino: {rutaCarpeta}")
        on_log(f"📄 Excel:  {rutaExcel}")
        on_log(f"🔢 Columna (1=A): {col}")
        on_log(f"🧠 Prompt: {gemi_txt}")
        on_log(f"🧠 Enviar a Gemini: {check_box}")
        
        rutaCarpeta.mkdir(parents=True, exist_ok=True)
        if rutaExcel.exists():
            wb = load_workbook(rutaExcel)
            on_log("✅ Excel abierto.")
        else:
            wb = Workbook()
            on_log("ℹ️  Excel no existía, se creó uno nuevo.")

        ws = wb.active
        ultima_fila = ws.max_row

        dest_resuelto = rutaCarpeta.resolve()

        copiadas = 0
        saltadas = 0
        errores = 0

        on_log("🚚 Copiando carpetas listadas en el Excel...")
        for fila in range(2, ultima_fila + 1):
            cancel_res = _check_cancel()
            if cancel_res:
                return cancel_res

            celda = ws.cell(row=fila, column=col)
            
            val = celda.value
            if val in (None, ""):
                continue

            origen = Path(str(val)).expanduser()
            try:
                origen_resuelto = origen.resolve(strict=False)
            except Exception:
                origen_resuelto = origen.absolute()

            if not origen.exists():
                on_log(f"⚠️  Fila {fila}: origen no existe -> {origen}")
                saltadas += 1
                continue

            if not origen.is_dir():
                on_log(f"⚠️  Fila {fila}: origen no es carpeta (es archivo) -> {origen}")
                saltadas += 1
                continue

            try:
                if origen_resuelto.is_relative_to(dest_resuelto):  # Python 3.9+
                    on_log(f"⚠️  Fila {fila}: origen está dentro del DESTINO. Se omite -> {origen}")
                    saltadas += 1
                    continue
            except AttributeError:
                if str(origen_resuelto).startswith(str(dest_resuelto)):
                    on_log(f"⚠️  Fila {fila}: origen está dentro del DESTINO. Se omite -> {origen}")
                    saltadas += 1
                    continue

            destino_final = rutaCarpeta / origen.name
            destino_final.parent.mkdir(parents=True, exist_ok=True)

            try:
                shutil.copytree(origen, destino_final, dirs_exist_ok=True)
                on_log(f"✅ Copiado: {origen}  ->  {destino_final}")
                copiadas += 1
            except Exception as e:
                on_log(f"❌ Error al copiar fila {fila} hacia {destino_final}: {e}")
                errores += 1

        ws.insert_cols(col)
        for fila in range(1, ws.max_row + 1):
            origen_cell = ws.cell(row=fila, column=col + 1)
            destino_cell = ws.cell(row=fila, column=col)
            if origen_cell.has_style:
                destino_cell.font = copy(origen_cell.font)
                destino_cell.border = copy(origen_cell.border)
                destino_cell.fill = copy(origen_cell.fill)
                destino_cell.number_format = copy(origen_cell.number_format)
                destino_cell.protection = copy(origen_cell.protection)
                destino_cell.alignment = copy(origen_cell.alignment)
        ws.cell(row=1, column=col, value="Observaciones")

        try:
            wb.save(ruta_excel)
            destino_archivo = Path(ruta_carpeta) / Path(ruta_excel).name
            shutil.copy2(ruta_excel, destino_archivo)
            on_log(f"💾 Copia del Excel en: {destino_archivo}")
        except PermissionError:
            on_log("❌ Revisa si el .xlsx está abierto en Excel o bloqueado por OneDrive/AV.")
        except Exception as e:
            on_log(f"❌ Error al copiar Excel: {e}")

        on_log(f"📊 Resumen copia: copiadas={copiadas}, saltadas={saltadas}, errores={errores}")

        if check_box is False:
            return {
                "copiadas": copiadas,
                "saltadas": saltadas,
                "errores": errores,
                "ruta_salida": None,
                "txt_creados": [],
            }

        on_log("\n🔍 Enviando archivos a Gemini para revisión...\n")

        salida_dir = rutaCarpeta / "respuestas_gemini"
        salida_dir.mkdir(parents=True, exist_ok=True)

        def _nombre_seguro(s: str) -> str:
            permitido = "-_.() abcdefghijklmnñopqrstuvwxyzABCDEFGHIJKLMNÑOPQRSTUVWXYZ0123456789"
            limpio = "".join(c if c in permitido else "_" for c in s)
            return "_".join(limpio.split())

        txt_creados = []

        for carpeta in rutaCarpeta.iterdir():
            cancel_res = _check_cancel()
            if cancel_res:
                return cancel_res

            if not carpeta.is_dir():
                continue
            if carpeta.resolve() == salida_dir.resolve():
                continue

            extensiones = ("*.txt", "*.pdf", "*.docx", "*.xlsx", "*.mp3")
            archivos_encontrados = []
            for ext in extensiones:
                archivos_encontrados += list(carpeta.glob(ext))

            if not archivos_encontrados:
                continue

            archivos = []
            for archivo in archivos_encontrados:
                cancel_res = _check_cancel()
                if cancel_res:
                    return cancel_res
                try:
                    if archivo.suffix.lower() == ".docx":
                        pdf_path = archivo.with_suffix(".pdf")
                        try:
                            convert(str(archivo), str(pdf_path))
                            archivo = pdf_path
                            on_log(f"📝 Convertido a PDF: {pdf_path.name}")
                        except Exception as e:
                            on_log(f"⚠️ Error convirtiendo {archivo.name} a PDF: {e}")
                            continue 

                    subido = genai.upload_file(str(archivo))
                    archivos.append(subido)
                    on_log(f"📤 Subido a Gemini: {archivo.name}")
                except Exception as e:
                    on_log(f"⚠️ Error subiendo {archivo.name}: {e}")

            if not archivos:
                continue

            try:
                prompt = gemi_txt
                respuesta = model.generate_content([prompt, *archivos])
                texto = f"📘 Carpeta/Obra: {carpeta.name}\n{respuesta.text}\n"

                nombre_txt = _nombre_seguro(carpeta.name) or "obra"
                ruta_txt = salida_dir / f"{nombre_txt}.txt"

                with open(ruta_txt, "w", encoding="utf-8") as f:
                    f.write(texto)

                txt_creados.append(str(ruta_txt))
                on_log(f"✅ Respuesta guardada: {ruta_txt}")
            except Exception as e:
                on_log(f"⚠️ Error generando/guardando respuesta para {carpeta.name}: {e}")

        if not txt_creados:
            on_log("\n⚠️ No hubo respuestas para guardar.")
            salida = None
        else:
            on_log(f"\n✅ {len(txt_creados)} archivos .txt creados en: {salida_dir}")
            salida = salida_dir

        return {
            "copiadas": copiadas,
            "saltadas": saltadas,
            "errores": errores,
            "ruta_salida": str(salida) if salida else None,
            "txt_creados": txt_creados,
        }

    except Exception as e:
        on_log(f"💥 Error fatal: {e}")
        return {"error": str(e)}
