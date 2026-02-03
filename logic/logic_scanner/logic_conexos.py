import re, os, unicodedata
import pdfplumber
from openpyxl import Workbook
from collections import OrderedDict

def normalizar_texto(texto: str) -> str:
    texto = texto.upper()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn") 
    texto = re.sub(r"[^A-Z0-9\s]", "", texto)  
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto

def normalizar_autores(texto: str) -> str:
    texto = texto.upper()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r",\s*(?=\S)", " / ", texto)
    texto = re.sub(r"[^A-Z0-9\s/]", "", texto)  
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto

def extraer_boletines_conexo(ruta):
    boletines = []

    for carpeta in os.listdir(ruta):
        ruta_sub = os.path.join(ruta, carpeta)

        if not os.path.isdir(ruta_sub):
            continue

        for f in os.listdir(ruta_sub):
            ruta_archivo = os.path.join(ruta_sub, f)

            if os.path.isfile(ruta_archivo) and re.match(r"^boletin_conexo", f, re.IGNORECASE):
                boletines.append(ruta_archivo)

    return boletines


def escanear_pdf_CM(carpetaBoletines, folio, stop_event=None):
    def check_cancel():
        return stop_event is not None and stop_event.is_set()

    boletines = extraer_boletines_conexo(carpetaBoletines)
    conexos = []
    tms = []

    for ruta_pdf in boletines:
        if check_cancel():
            return {"cancelado": True}

        with pdfplumber.open(ruta_pdf) as pdf:
            tablas = []
            temas = []
            tema_actual = None

            texto_partes = []

            for page in pdf.pages:
                if check_cancel():
                    return {"cancelado": True}

                texto_partes.append(page.extract_text() or "")
                tablas.extend(page.extract_tables() or [])

            texto = "\n".join(texto_partes)

            m = re.search(r"Año de Publicación\s*:\s*(\d{4})", texto)
            anioPubCon = int(m.group(1)) if m else None

            m1 = re.search(r"N°\s*Boletín\s*:\s*(\d+)", texto)
            numBoletin = m1.group(1) if m1 else None
            
            if len(tablas) < 2 or not tablas[0] or not tablas[1]:
                continue

            cabeceraInfoCon = tablas[0][0]
            filaInfoCon = tablas[1][0]
            infoCon = dict(zip(cabeceraInfoCon, filaInfoCon))
            
            tm_actual = {"Titulares_master": []}

            for tabla in tablas:
                if check_cancel():
                    return {"cancelado": True}

                for fila in tabla:
                    if not fila or not fila[0]:
                        continue

                    key = str(fila[0]).strip().lower()
                    val = fila[1] if len(fila) > 1 else None

                    if key == "titulo":
                        tema_actual = {
                            "Titulo": val,
                            "Autor": None,
                            "Interpretes_principales": []
                        }
                        temas.append(tema_actual)
                        continue

                    if not tema_actual:
                        continue

                    if key == "autor":
                        tema_actual["Autor"] = val
                        continue

                if tabla and tabla[0] and str(tabla[0][0]).strip().lower() == "interpretes principales":
                    if tema_actual:
                        tema_actual["Interpretes_principales"] = extraer_datos_dh(tabla, tm=False)
                        
                if tabla and tabla[0] and str(tabla[0][0]).strip().lower() == "interpretes ejecutantes":
                    if tema_actual:
                        tema_actual["Interpretes_ejecutantes"] = extraer_datos_dh(tabla, incluir_tipo_part=True, tm=False)

                if tabla and tabla[0] and str(tabla[0][0]).strip().lower() == "titular master":
                    if tm_actual:
                        tm_actual["Titulares_master"] = extraer_datos_dh(tabla, tm=True)
                        
            conexoLabelCopy = {
                "Declarante": infoCon.get("Declarante", ""),
                "Nombre Disco": infoCon.get("Nombre Disco", ""),
                "Artista Principal o grupo": infoCon.get("Nombre Artista o Grupo", ""),
                "Fecha Publ": anioPubCon,
                "Temas": temas,
            }
            
            conexoReparto = {
                "Bol": numBoletin,
                "Nombre Disco": infoCon.get("Nombre Disco", ""),
                "Temas": temas
            }
            
            tm = {
                "Bol": numBoletin,
                "Titulares_master": tm_actual.get("Titulares_master", [])
            }
            
            tms.append(tm)
            
        conexo = {
            "label": conexoLabelCopy,
            "reparto": conexoReparto
        }
        
        conexos.append(conexo)

    LC = exportar_excel_LabelCopy(conexos, folio, stop_event=stop_event)
    R = exportar_excel_reparto(conexos, folio, stop_event=stop_event)
    TM = exportar_excel_tm(tms, folio, stop_event=stop_event)
    
    if LC is True and R is True and TM is True:
        return True

def parse_linea(linea):
    if isinstance(linea, list):
        linea = linea[0]
    if not linea:
        return []
    return str(linea).split()

def nombre_completo(i: dict) -> str:
    return " ".join(filter(None, [
        i.get("nombre", "").replace("\n", " ").strip(),
        i.get("apellido_paterno", "").strip(),
        i.get("apellido_materno", "").strip(),
    ])).strip()
    
def extraer_datos_dh(tabla, incluir_tipo_part=False, tm=False):
    dhs = []
    for fila in tabla[2:]:
        if fila and len(fila) > 6 and fila[3] and not tm:
            interprete = {
                "numero": fila[0],
                "nombre": fila[3],
                "apellido_paterno": fila[4],
                "apellido_materno": fila[5],
                "ipi": fila[6],
            }

            if incluir_tipo_part and len(fila) > 7:
                interprete["Tipo Part"] = fila[7]
                
            dhs.append(interprete)
            
        if fila and len(fila) > 7 and tm:
            titulares_master = {
                "numero": fila[0],
                "nombre": fila[3],
                "apellido_paterno": fila[4],
                "apellido_materno": fila[5],
                "porcentaje": fila[6],
                "ipi": fila[7]
            }
            
            dhs.append(titulares_master)
            
    return dhs

def exportar_excel_tm(boletines, folio, stop_event=None):
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    wb = Workbook()
    ws = wb.active
    ws.title = "hoja1"

    ws["A1"] = "Boletín"
    ws["B1"] = "Nombre DH Master"
    ws["C1"] = "IPI"
    ws["D1"] = "PART"

    fila = 2

    for boletin in boletines:
        if check_cancel():
            return {"cancelado": True}
        bol = boletin.get("Bol", "")
        titulares = boletin.get("Titulares_master", [])
        
        ws.cell(row=fila, column=1, value=bol)
    
        for titular in titulares:
            if check_cancel():
                return {"cancelado": True}
            
            nombre = " ".join(nombre_completo(titular).upper().split())
            porcentaje = titular.get("porcentaje", "")
            ipi = titular.get("ipi", "")
            
            ws.cell(row=fila, column=2, value=normalizar_autores(nombre))
            ws.cell(row=fila, column=3, value=ipi)
            ws.cell(row=fila, column=4, value=porcentaje)
            
            fila += 1
            
    escritorio = os.path.join(
        os.path.expanduser("~"),
        r"OneDrive - Sociedad Chilena de Autores e Interpretes Musicales\Escritorio"
    )
    os.makedirs(escritorio, exist_ok=True)

    wb.save(os.path.join(escritorio, f"CMC_TM_{folio}.xlsx"))
    return True

def exportar_excel_reparto(boletines, folio, stop_event=None):
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    wb = Workbook()
    ws = wb.active
    ws.title = "hoja1"
    
    ws["A1"] = "N° Bol"
    ws["B1"] = "Title Producto / titulo fonograma"
    ws["C1"] = "Song Title / Título Obra"
    ws["D1"] = "Nombre DH"
    ws["E1"] = "IP"
    ws["F1"] = "IS"
    ws["G1"] = "IPI"
    ws["H1"] = "GID"
    
    fila = 2

    for boletin in boletines:
        if check_cancel():
            return {"cancelado": True}
        reparto = boletin.get("reparto", {}) or {}
        bol = reparto.get("Bol")

        if isinstance(bol, list):
            bol = bol[0] if bol else None

        nombre_disco = reparto.get("Nombre Disco", "")
        temas = reparto.get("Temas", []) or []

        ws.cell(row=fila, column=2, value=normalizar_texto(nombre_disco))  # B
        ws.cell(row=fila, column=1, value=bol)
        
        
        for tema in temas:
            if check_cancel():
                return {"cancelado": True}
            titulo = tema.get("Titulo", "")

            # --- Fila cabecera del tema (A,B,C) ---
            ws.cell(row=fila, column=3, value=normalizar_texto(titulo))
            ws.cell(row=fila, column=8, value="C")   
        
            # --- Filas de intérpretes (D,E/F,G) ---
            interpretesE = tema.get("Interpretes_ejecutantes", []) or []
            interpretesP = tema.get("Interpretes_principales", []) or []

            final = OrderedDict()  
            
            def key_it(it):
                nombre = " ".join(nombre_completo(it).upper().split())
                ipi = it.get("ipi", "") or it.get("IPI", "") or ""
                return (nombre, ipi)

            # 1) Principales -> IP = 1
            for it in interpretesP:
                k = key_it(it)
                if k not in final:
                    final[k] = {"ip": 0, "is": 0}
                final[k]["ip"] = 1

            # 2) Ejecutantes -> IS suma Tipo Part
            for it in interpretesE:
                k = key_it(it)
                if k not in final:
                    final[k] = {"ip": 0, "is": 0}

                tipo_raw = (it.get("Tipo Part") or "").strip()
                tipo = int(tipo_raw) if tipo_raw.isdigit() else 0
                final[k]["is"] += tipo

            # 3) Escribir una sola vez
            for (nombre, ipi), tot in final.items():
                ws.cell(row=fila, column=4, value=normalizar_autores(nombre))  
                if tot["ip"] == 1:
                    ws.cell(row=fila, column=5, value=1)                      
                if tot["is"]:
                    ws.cell(row=fila, column=6, value=tot["is"])               
                if ipi:
                    ws.cell(row=fila, column=7, value=ipi)                     
                fila += 1
                
            fila += 1
            
    escritorio = os.path.join(
        os.path.expanduser("~"),
        r"OneDrive - Sociedad Chilena de Autores e Interpretes Musicales\Escritorio"
    )
    os.makedirs(escritorio, exist_ok=True)

    wb.save(os.path.join(escritorio, f"CMC_R_{folio}.xlsx"))
    return True
    
def exportar_excel_LabelCopy(boletines, folio, stop_event=None):
    def check_cancel():
        return stop_event is not None and stop_event.is_set()

    wb = Workbook()
    ws = wb.active
    ws.title = "hoja1"

    ws["A1"] = "Folio"
    ws["C1"] = "Title Producto / titulo fonograma"
    ws["D1"] = "Declarante"
    ws["E1"] = "Tune Key / Codigo Producto"
    ws["G1"] = "Artist Producto / Artista Principal"
    ws["H1"] = "Soporte / tipo producto"
    ws["I1"] = "Marca"
    ws["K1"] = "Numero Lado"
    ws["L1"] = "Tune Number / N° Track"
    ws["M1"] = "Fecha Publ"
    ws["N1"] = "Song Title / Título Obra"
    ws["P1"] = "Song Author / Autores"
    ws["Q1"] = "Artist / Interprete Obra"
    ws["U1"] = "ISRC"

    fila_excel = 2
    contador = 1
    
    for boletin in boletines:
        if check_cancel():
            return {"cancelado": True}
        
        label = boletin.get("label", {}) or {}
        temas = label.get("Temas") or []

        for n_track, tema in enumerate(temas, start=1):
            if check_cancel():
                return {"cancelado": True}
            
            ws.cell(row=fila_excel, column=1, value=folio)  
            ws.cell(row=fila_excel, column=3, value=normalizar_texto(label.get("Nombre Disco")))
            ws.cell(row=fila_excel, column=4, value=normalizar_texto(label.get("Declarante")))   
            ws.cell(row=fila_excel, column=5, value=f"SCD-{folio}-{contador}")     
            ws.cell(row=fila_excel, column=7, value=normalizar_texto(label.get("Artista Principal o grupo"))) 
            ws.cell(row=fila_excel, column=11, value="1") 
            ws.cell(row=fila_excel, column=12, value=n_track)  
            ws.cell(row=fila_excel, column=13, value=f"01/{label.get('Fecha Publ')}")        
            ws.cell(row=fila_excel, column=17, value=normalizar_texto(label.get("Artista Principal o grupo")))

            # ---- Datos del tema  ----
            ws.cell(row=fila_excel, column=14, value=normalizar_texto(tema.get("Titulo")))
            ws.cell(row=fila_excel, column=16, value=normalizar_autores(tema.get("Autor")))
            fila_excel += 1
            contador += 1
            
    escritorio = os.path.join(
        os.path.expanduser("~"),
        r"OneDrive - Sociedad Chilena de Autores e Interpretes Musicales\Escritorio"
    )
    os.makedirs(escritorio, exist_ok=True)

    wb.save(os.path.join(escritorio, f"CMC_LC_{folio}.xlsx"))
    return True
