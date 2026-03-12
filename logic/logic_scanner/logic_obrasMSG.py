import os, re, csv, unicodedata
from openpyxl import load_workbook
from collections import defaultdict, OrderedDict
from utils.config_manager import obtener_carpeta_base

EXTS_EXCEL = (".xlsx", ".xlsm", ".xltx", ".xltm")

# =========================
# UTIL
# =========================
def extraer_excels(ruta):
    return [
        os.path.join(ruta, f)
        for f in os.listdir(ruta)
        if os.path.isfile(os.path.join(ruta, f)) and f.lower().endswith(EXTS_EXCEL)
    ]

def hoja_a_tabla(ws, max_rows=None, max_cols=None):
    rows = []
    for i, row in enumerate(ws.values, start=1):
        if max_rows and i > max_rows:
            break
        if max_cols:
            row = row[:max_cols]
        rows.append(list(row))
    while rows and all(v is None or str(v).strip() == "" for v in rows[-1]):
        rows.pop()
    return rows

def tabla_a_texto(tabla):
    lineas = []
    for row in tabla:
        celdas = [str(v).strip() for v in row if v is not None and str(v).strip() != ""]
        if celdas:
            lineas.append(" ".join(celdas))
    return "\n".join(lineas)

def parse_ipi(p):
    return (p or "").lstrip("0")

def parse_name(titulo: str) -> str:
    return re.sub(r"\s*\(.*?\)", "", (titulo or "")).strip()

def normalizar_nombre(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto or "")
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^A-Za-z\s]", "", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto

def revisar_porcentajes(obras: dict):
    alertas = []

    for key, obra in obras.items():
        titulo = obra.get("titulo") or "(sin título)"
        iswc = obra.get("iswc") or "(sin ISWC)"

        suma_p = sum(r.get("p_share", 0.0) for r in obra["dh"])
        suma_m = sum(r.get("m_share", 0.0) for r in obra["dh"])

        suma_p = round(suma_p, 2)
        suma_m = round(suma_m, 2)
        
        if suma_p != 100 or (suma_m != 100 and suma_m > 0):
            alertas.append((key, titulo, iswc, suma_p, suma_m))

    return alertas

# ============================================================
# MODO 1 
# ============================================================
LINE_RE = re.compile(
    r'^(?P<nombre>.+?)\s+'
    r'(?P<rol>[A-Z]{1,3})\s+'
    r'(?P<ipn>\d{8,12})\s+'
    r'\d+\s+\d+\s+'
    r'(?P<p_share>\d+(?:[.,]\d+)?)\s+'
    r'\d+\s+'
    r'(?P<m_share>\d+(?:[.,]\d+)?)\s*$'
)

def extraer_titulo_afitap(text: str) -> str | None:
    lines = [ln.strip() for ln in text.splitlines()]
    for i, ln in enumerate(lines):
        if re.match(r'^---\s*HOJA:\s*Report\s*---\s*$', ln, re.IGNORECASE):
            for j in range(i + 1, min(i + 10, len(lines))):
                if lines[j]:
                    return lines[j]
    return None

def parse_dh_line(line: str) -> dict | None:
    m = LINE_RE.match(line.strip())
    if not m:
        return None

    d = m.groupdict()
    d["p_share"] = float(str(d["p_share"]).replace(",", "."))

    if d["m_share"]:
        d["m_share"] = float(str(d["m_share"]).replace(",", "."))
    else:
        d["m_share"] = 0.0
        d["m_soc"] = None

    return d

def parse_lines(text: str):
    rows = []
    current_title = extraer_titulo_afitap(text)

    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue

        d = parse_dh_line(line)
        if not d or not current_title:
            continue

        d["nombre_obra"] = current_title
        rows.append(d)

    return rows

def agrupar_por_titulo(rows):
    obras = defaultdict(lambda: {"titulo": None, "dh": []})

    for r in rows:
        titulo = r["nombre_obra"]
        obras[titulo]["titulo"] = titulo
        obras[titulo]["dh"].append(r)

    return dict(obras)

def consolidar_dh_por_ipn(obras: dict) -> dict:
    for titulo, obra in obras.items():
        acumulado = OrderedDict()

        for r in obra["dh"]:
            key = (r["ipn"], r["rol"])

            if key not in acumulado:
                acumulado[key] = dict(r)
            else:
                acumulado[key]["p_share"] += r["p_share"]
                acumulado[key]["m_share"] += r["m_share"]

        obra["dh"] = list(acumulado.values())

    return obras

def scanner_excel_modo_1(ruta, stop_event=None):
    
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    
    rutas = extraer_excels(ruta)
    all_rows_global = []

    for ruta in rutas:
        if check_cancel():
                return {"cancelado": True}
        try:
            wb = load_workbook(ruta, data_only=True)
        except Exception as e:
            print(f"⚠️ No pude abrir: {ruta}\n   Error: {e}")
            continue

        for nombre_hoja in wb.sheetnames:
            if check_cancel():
                return {"cancelado": True}
            ws = wb[nombre_hoja]
            tabla = hoja_a_tabla(ws)
            if not tabla:
                continue

            texto = f"--- HOJA: {nombre_hoja} ---\n" + tabla_a_texto(tabla)
            rows = parse_lines(texto)
            all_rows_global.extend(rows)

    obras = agrupar_por_titulo(all_rows_global)
    obras = consolidar_dh_por_ipn(obras)
    return obras

# ============================================================
# MODO 2 
# ============================================================
def to_float(v):
    if v is None:
        return 0.0
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except:
        return 0.0

def norm_nombre_key(s: str) -> str:
    s = (s or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Za-z\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip().upper()
    return s

def leer_excel_por_headers(path_xlsx):
    wb = load_workbook(path_xlsx, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    req = ["WorkNumber","WorkOriginalTitle","RoleCode","Name","IPNameNr","PRShareCollection","MRShareCollection"]
    faltan = [h for h in req if h not in idx]
    if faltan:
        raise ValueError(f"{os.path.basename(path_xlsx)}: faltan columnas {faltan}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        work = row[idx["WorkNumber"]]
        if not work:
            continue

        rows.append({
            "work_number": str(work).strip(),
            "titulo": str(row[idx["WorkOriginalTitle"]] or "").strip(),
            "rol": str(row[idx["RoleCode"]] or "").strip(),
            "nombre": str(row[idx["Name"]] or "").strip(),
            "ipn": str(row[idx["IPBaseNr"]] or "").strip(),
            "p_share": to_float(row[idx["PRShareCollection"]]),
            "m_share": to_float(row[idx["MRShareCollection"]]),
            "iswc": str(row[idx.get("WorkISWC","")] if "WorkISWC" in idx else "").strip(),
        })
    return rows

def agrupar_por_worknumber(rows):
    obras = defaultdict(lambda: {"titulo": None, "work_number": None, "dh": []})
    for r in rows:
        key = r["work_number"]
        obras[key]["titulo"] = r["titulo"]
        obras[key]["work_number"] = key
        obras[key]["dh"].append(r)
    return dict(obras)

def consolidar_dh_modo_2(obras):
    for _, obra in obras.items():
        acc = OrderedDict()

        for r in obra["dh"]:
            ipn = (r.get("ipn") or "").strip()

            if ipn and set(ipn) != {"0"}:
                key = ("IPI", ipn, r.get("rol"))
            else:
                key = ("NOMBRE", r.get("rol"), norm_nombre_key(r.get("nombre")))

            if key not in acc:
                acc[key] = dict(r)
                if not ipn or set(ipn) == {"0"}:
                    acc[key]["ipn"] = ""
            else:
                acc[key]["p_share"] += r.get("p_share", 0.0)
                acc[key]["m_share"] += r.get("m_share", 0.0)

        obra["dh"] = list(acc.values())

    return obras

def scanner_excel_modo_2(ruta, stop_event=None):
    
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    
    rutas = extraer_excels(ruta)
    rows_global = []

    for ruta in rutas:
        if check_cancel():
                return {"cancelado": True}
        try:
            rows_global.extend(leer_excel_por_headers(ruta))
        except Exception as e:
            print(f"⚠️ Saltando {ruta}: {e}")

    obras = agrupar_por_worknumber(rows_global)
    obras = consolidar_dh_modo_2(obras)
    return obras

# ============================================================
# EXPORT 
# ============================================================
def export_excel(obras, folio, stop_event=None):
    
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    
    carpeta_base = obtener_carpeta_base()
    
    ruta_destino = carpeta_base / "MMs" / f"mm_msg_{folio}.csv"
    
    with open(
        ruta_destino,
        "w",
        newline="",
        encoding="latin-1"
    ) as f:
        writer = csv.writer(f, delimiter=";")

        for _, obra in obras.items():
            if check_cancel():
                return {"cancelado": True}
            writer.writerow([parse_name(obra.get("titulo")), "", "", "", "", ""])
            writer.writerow([])
            writer.writerow([])
            writer.writerow([])

            for r in obra["dh"]:
                if check_cancel():
                    return {"cancelado": True}
                writer.writerow([
                    r.get("rol", ""),
                    normalizar_nombre(r.get("nombre", "")),
                    f"{float(r.get('p_share', 0.0)):.2f}".replace(".", ","),
                    f"{float(r.get('m_share', 0.0)):.2f}".replace(".", ","),
                    "",
                    parse_ipi(r.get("ipn", "")),
                ])

            writer.writerow([])
    
    return {"ok": True, "ruta_destino": ruta_destino}

def scanner_excel(ruta, folio, stop_event=None, modo=1):
    
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    
    if modo == 1:
        if check_cancel():
                return {"cancelado": True}
        obras = scanner_excel_modo_1(ruta, stop_event=stop_event)
        alertas = revisar_porcentajes(obras)
    else:
        if check_cancel():
                return {"cancelado": True}
        obras = scanner_excel_modo_2(ruta, stop_event=stop_event)
        alertas = revisar_porcentajes(obras)
        
    res = export_excel(obras, folio, stop_event=stop_event)
    
    if res.get("ok"):
        return {"ok": True, "cantidad_obras": len(obras), "ruta_destino": res.get("ruta_destino"), "alertas": alertas}

