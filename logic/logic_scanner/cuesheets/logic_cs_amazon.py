import pdfplumber, re, csv, unicodedata
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from utils.config_manager import obtener_carpeta_base

# =========================
# Normalización
# =========================
def normalizar_nombre(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^A-Za-z0-9\s\+\-]", "", texto)  # permite + y -
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto.upper()

# =========================
# Parsers
# =========================

DH_RE = re.compile(
    r'^\s*'
    r'(?P<rol>[A-Z]{1,3})\s+'
    r'(?P<nombre>.+?)\s+'
    r'(?P<soc_dh>[A-Za-z\/\-]+)\s+'
    r'(?P<pct>\d+(?:\.\d+)?)\s*$'
)

TITLE_RE = re.compile(
    r'^\s*(?:\d+\s+)*'          # basura: "5 5 " (cualquier cantidad de números)
    r'(?P<title>.+?)\s+'        # título
    r'(?P<soc_obra>[A-Za-z]{1,5})\s+'# sociedad
    r'(?P<min>\d{1,2}):(?P<sec>\d{2})\s*$'
)

EPISODE_TITLE = re.compile(
    r'^Episode Title:\s*"(?P<ep_title>.+?)"'
)

def extraer_episode_title(texto: str) -> str:
    for raw in texto.splitlines():
        m = EPISODE_TITLE.match(raw.strip())
        if m:
            return m.group("ep_title").strip()
    return ""

def parse_dh_line(line: str) -> dict | None:
    m = DH_RE.match(line.strip())
    if not m:
        return None

    d = m.groupdict()
    
    d["nombre"] = normalizar_nombre(d["nombre"])

    d["pct"] = float(d["pct"].replace(",", "."))

    return d

def parse_lines(text: str):
    rows = []
    current = None  # dict con metadata del título actual

    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue

        tm = TITLE_RE.match(line)
        if tm:
            current = {
                "nombre_obra": tm.group("title").strip(),
                "soc_obra": tm.group("soc_obra").strip().upper(),
                "min": int(tm.group("min")),
                "sec": int(tm.group("sec")),
                "_obra_key": f"NOMBRE:{normalizar_nombre(tm.group('title'))}|SOC:{tm.group('soc_obra').strip().upper()}|T:{tm.group('min')}:{tm.group('sec')}"
            }
            continue

        d = parse_dh_line(line)
        if not d or not current:
            continue

        # pega metadata a cada DH
        d["nombre_obra"] = current["nombre_obra"]
        d["soc_obra"] = current["soc_obra"]
        d["min"] = current["min"]
        d["sec"] = current["sec"]
        d["_obra_key"] = current["_obra_key"]

        rows.append(d)

    return rows

# =========================
# Agrupar + consolidar
# =========================

def agrupar_por_obra(rows):
    obras = defaultdict(lambda: {"titulo": None, "soc_obra": None, "min": None, "sec": None, "dh": []})

    for r in rows:
        key = r["_obra_key"]
        obra = obras[key]
        obra["titulo"] = r.get("nombre_obra")
        obra["soc_obra"] = r.get("soc_obra")
        obra["min"] = r.get("min")
        obra["sec"] = r.get("sec")
        obra["dh"].append(r)

    return dict(obras)

def _keypart(x):
    if x is None:
        return ""
    if isinstance(x, list):
        return "|".join(str(i).strip() for i in x)
    return str(x).strip()

def consolidar_dh_por_ipn(obras: dict) -> dict:
    for key, obra in obras.items():
        acumulado = OrderedDict()

        for r in obra["dh"]:
            k = (_keypart(r.get("rol")), _keypart(r.get("nombre")))


            if k not in acumulado:
                acumulado[k] = dict(r)
            else:
                acumulado[k]["pct"] += r["pct"]

        obra["dh"] = list(acumulado.values())

    return obras

def compositores_str(dh_list):
    cs = [r for r in dh_list if r.get("rol") != "E"]

    cs.sort(key=lambda r: float(r.get("pct", 0.0)))  # ordenar por %

    nombres = [r.get("nombre", "").strip() for r in cs if r.get("nombre")]

    return "/".join(nombres)

def revisar_porcentajes(obras: dict):
    alertas = []

    for key, obra in obras.items():
        titulo = obra.get("titulo") or "(sin título)"
        iswc = obra.get("iswc") or "(sin ISWC)"

        suma_p = sum(r.get("p_share", 0.0) for r in obra["dh"])
        suma_m = sum(r.get("m_share", 0.0) for r in obra["dh"])

        suma_p = round(suma_p, 2)
        suma_m = round(suma_m, 2)

        if suma_p != 100 or suma_m != 100:
            alertas.append((key, titulo, iswc, suma_p, suma_m))

    return alertas

def parse_ipi(p):
    if p is None or str(p).strip() == "":
        return "SIN_IPI"   # o "" si no quieres texto
    
    return str(p).strip().lstrip("0")

def parse_name(titulo: str) -> str:
    titulo = titulo.replace("ISWC:", "")

    titulo = titulo.replace("(", " ").replace(")", " ")

    titulo = unicodedata.normalize("NFKD", titulo)
    titulo = "".join(c for c in titulo if not unicodedata.combining(c))

    titulo = titulo.replace("&", " Y ")

    titulo = re.sub(r"[^A-Za-z0-9\s]", "", titulo)

    titulo = re.sub(r"\s+", " ", titulo).strip()

    return titulo.upper()

# =========================
# Export
# =========================

def export_excel(obras, episode_title: str, stop_event=None, datos: dict = None):
    
    fila_por_obra = {}
    
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    
    folio = datos.get("folio", "")
    cap = datos.get("cap", "")
    
    carpeta_base = obtener_carpeta_base()
    
    ruta_destino = carpeta_base / "MMs" / f"cs_amazonNAC_{folio}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "hoja1"

    # 1) Fila 1: episodio SOLO
    ws.cell(row=1, column=1, value=int(datos.get("folio", "")))
    ws.cell(row=1, column=2, value=3)  # val=1 solo en la primera línea
    ws.cell(row=1, column=3, value=normalizar_nombre(episode_title))
    ws.cell(row=1, column=4, value=f"CAP {cap}")
    ws.cell(row=1, column=7, value=normalizar_nombre(datos.get("dir", "")))
    antag = normalizar_nombre(datos.get("antog", ""))
    protag = normalizar_nombre(datos.get("protag", ""))
    ws.cell(row=1, column=8, value=f"{protag}/{antag}")
    ws.cell(row=1, column=11, value=int(datos.get("anio", "")))

    # 2) Desde fila 2: episodio + obra alineados
    fila = 2
    for key, obra in obras.items():
        fila_por_obra[key] = fila

        # episodio (repetido en cada fila desde la 2, con val=0)
        ws.cell(row=fila, column=1, value=int(datos.get("folio", "")))
        ws.cell(row=fila, column=2, value=0)
        ws.cell(row=fila, column=3, value=normalizar_nombre(episode_title))
        ws.cell(row=fila, column=4, value=f"CAP {cap}")
        ws.cell(row=fila, column=7, value=normalizar_nombre(datos.get("dir", "")))
        ws.cell(row=fila, column=8, value=f"{antag}/{protag}")
        ws.cell(row=fila, column=11, value=int(datos.get("anio", "")))

        # obra (empieza en fila 2)
        ws.cell(row=fila, column=12, value=parse_name(obra["titulo"]))
        ws.cell(row=fila, column=13, value=obra.get("soc_obra", ""))
        ws.cell(row=fila, column=14, value=1)
        ws.cell(row=fila, column=15, value=int(obra.get("min", 0)))
        ws.cell(row=fila, column=16, value=int(obra.get("sec", 0)))
        ws.cell(row=fila, column=17, value=compositores_str(obra["dh"]))

        fila += 1

    wb.save(ruta_destino)
    return {"fila_por_obra": fila_por_obra, "ruta_destino": ruta_destino}

# =========================
# Scanner
# =========================

def scanner(ruta, datos, stop_event=None):
    
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    
    with pdfplumber.open(ruta) as pdf:
        if check_cancel():
            return {"cancelado": True}
        texto = "\n".join(page.extract_text() or "" for page in pdf.pages)

    episode_title = extraer_episode_title(texto)
    rows = parse_lines(texto)
    obras = agrupar_por_obra(rows)
    obras = consolidar_dh_por_ipn(obras)
    
    res = export_excel(obras, episode_title, stop_event=stop_event, datos=datos)

    cantidad_obras = len(obras)
    
    return {"cantidad_obras": cantidad_obras, "ok": True, "ruta_destino": res.get("ruta_destino", "desconocida")}
    
