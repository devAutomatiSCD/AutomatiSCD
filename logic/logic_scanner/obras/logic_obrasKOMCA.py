import os, re, csv, unicodedata, xlrd
from collections import OrderedDict
from openpyxl import load_workbook
from utils.config_manager import obtener_carpeta_base

EXTS_EXCEL = (".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")

def limpiar(valor):
    if valor is None:
        return ""
    return str(valor).strip()

def es_excel(ruta):
    return ruta.lower().endswith(EXTS_EXCEL)

def a_float(valor):
    valor = limpiar(valor)
    if not valor:
        return 0.0

    valor = valor.replace("%", "").strip()
    valor = valor.replace(",", ".")

    try:
        return float(valor)
    except ValueError:
        return 0.0

# =========================
# Adaptadores de lectura
# =========================

class OpenPyxlSheetAdapter:
    def __init__(self, ws):
        self.ws = ws
        self.max_row = ws.max_row

    def cell_value(self, fila, col):
        return self.ws.cell(fila, col).value


class XlrdSheetAdapter:
    def __init__(self, ws):
        self.ws = ws
        self.max_row = ws.nrows

    def cell_value(self, fila, col):
        try:
            return self.ws.cell_value(fila - 1, col - 1)
        except IndexError:
            return None


def abrir_excel(ruta_excel, hoja=None):
    ext = os.path.splitext(ruta_excel)[1].lower()

    if ext not in EXTS_EXCEL:
        raise ValueError(f"Extensión no soportada: {ext}")

    if ext == ".xls":
        wb = xlrd.open_workbook(ruta_excel)
        ws = wb.sheet_by_name(hoja) if hoja else wb.sheet_by_index(0)
        return XlrdSheetAdapter(ws)

    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb[hoja] if hoja else wb.active
    return OpenPyxlSheetAdapter(ws)


def es_fila_vacia(ws, fila, max_col=10):
    for col in range(1, max_col + 1):
        if limpiar(ws.cell_value(fila, col)):
            return False
    return True


def leer_obras_excel(ruta_excel, hoja=None, stop_event=None):
    
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    
    if not es_excel(ruta_excel):
        raise ValueError(f"El archivo no es Excel soportado: {ruta_excel}")

    ws = abrir_excel(ruta_excel, hoja=hoja)

    obras = {}
    fila = 1
    max_fila = ws.max_row

    while fila <= max_fila:
        if check_cancel():
            return {"cancelado": True}
        etiqueta = limpiar(ws.cell_value(fila, 1)).upper()

        if etiqueta == "TITLE":
            obra = {
                "titulo": limpiar(ws.cell_value(fila, 3)),
                "sub_titulo": None,
                "iswc": None,
                "dh": []
            }

            fila += 1

            while fila <= max_fila:
                if check_cancel():
                    return {"cancelado": True}
                col_a = limpiar(ws.cell_value(fila, 1)).upper()

                if col_a == "SUB TITLE":
                    obra["sub_titulo"] = limpiar(ws.cell_value(fila, 3))

                elif col_a == "ISWC":
                    obra["iswc"] = limpiar(ws.cell_value(fila, 3))

                elif col_a == "CAE":
                    fila += 1

                    while fila <= max_fila:
                        if check_cancel():
                            return {"cancelado": True}
                        cae = limpiar(ws.cell_value(fila, 1))
                        nombre = limpiar(ws.cell_value(fila, 2))
                        ipi = limpiar(ws.cell_value(fila, 4))
                        p_soc = limpiar(ws.cell_value(fila, 6))
                        p_part = limpiar(ws.cell_value(fila, 7))
                        m_soc = limpiar(ws.cell_value(fila, 8))
                        m_part = limpiar(ws.cell_value(fila, 9))

                        if not any([cae, nombre, ipi, p_soc, p_part, m_soc, m_part]):
                            break

                        texto_fila = " ".join(
                            limpiar(ws.cell_value(fila, c)) for c in range(1, 5)
                        ).upper()

                        if "KOMCA CODE" in texto_fila or "ARTIST" in texto_fila or "TITLE" in texto_fila:
                            break

                        obra["dh"].append({
                            "rol": cae,
                            "nombre": nombre,
                            "ipi": ipi,
                            "p_soc": p_soc,
                            "p_share": a_float(p_part),
                            "m_soc": m_soc,
                            "m_share": a_float(m_part),
                        })

                        fila += 1

                    continue

                elif col_a == "TITLE":
                    break

                fila += 1

            key = obra["iswc"] or f"sin_iswc_{len(obras) + 1}"

            if key in obras:
                if check_cancel():
                    return {"cancelado": True}
                suf = 2
                new_key = f"{key}_{suf}"
                while new_key in obras:
                    suf += 1
                    new_key = f"{key}_{suf}"
                key = new_key

            obras[key] = obra
            continue

        fila += 1

    return obras


def consolidar_dh_por_ipn(obras: dict) -> dict:
    for iswc, obra in obras.items():
        acumulado = OrderedDict()

        for r in obra["dh"]:
            key = (r.get("ipi", ""), r.get("rol", ""))

            p = float(r.get("p_share", 0) or 0)
            m = float(r.get("m_share", 0) or 0)

            if key not in acumulado:
                nuevo = dict(r)
                nuevo["p_share"] = p
                nuevo["m_share"] = m
                acumulado[key] = nuevo
            else:
                acumulado[key]["p_share"] += p
                acumulado[key]["m_share"] += m

        for reg in acumulado.values():
            reg["p_share"] = round(reg["p_share"], 2)
            reg["m_share"] = round(reg["m_share"], 2)

        obra["dh"] = list(acumulado.values())

    return obras


def revisar_porcentajes(obras: dict):
    alertas = []

    for iswc, obra in obras.items():
        titulo = obra.get("titulo") or "(sin título)"

        suma_p = sum(float(r.get("p_share", 0) or 0) for r in obra["dh"])
        suma_m = sum(float(r.get("m_share", 0) or 0) for r in obra["dh"])

        suma_p = round(suma_p, 2)
        suma_m = round(suma_m, 2)

        if suma_p != 100 or (suma_m != 100 and suma_m > 0):
            alertas.append((iswc, titulo, suma_p, suma_m))

    return alertas


def parse_ipi(p):
    return limpiar(p).lstrip("0")


def parse_name(titulo: str) -> str:
    return re.sub(r"\s*\(.*?\)", "", limpiar(titulo)).strip()


def normalizar_nombre(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^A-Za-z\s]", "", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def export_excel(obras, folio, stop_event=None):
    
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    
    carpeta_base = obtener_carpeta_base()
    
    ruta_destino = carpeta_base / "MMs" / f"mm_komca_{folio}.csv"
    
    fila_por_obra = {}
    fila = 1
    
    with open(
        ruta_destino,
        "w",
        newline="",
        encoding="latin-1"
    ) as f:
        writer = csv.writer(f, delimiter=";")

        for key, obra in obras.items():
            if check_cancel():
                return {"cancelado": True}
            fila_por_obra[key] = fila
            writer.writerow([parse_name(obra["titulo"]), "", "", "", "", ""])
            writer.writerow([])
            writer.writerow([])
            writer.writerow([])

            suma_m_obra = round(sum((x.get("m_share") or 0.0) for x in obra["dh"]), 2)
            mostrar_m = suma_m_obra > 0

            for r in obra["dh"]:
                if check_cancel():
                    return {"cancelado": True}
                m_cell = (
                    f"{r['m_share']:.2f}".replace(".", ",")
                    if mostrar_m
                    else ""
                )

                writer.writerow([
                    r["rol"],
                    normalizar_nombre(r["nombre"]),
                    f"{r['p_share']:.2f}".replace(".", ","),
                    m_cell,
                    "",
                    parse_ipi(r["ipi"]),
                ])

            writer.writerow([])

    return {"fila_por_obra": fila_por_obra, "ok": True, "ruta_destino": ruta_destino}

def scanner(ruta, folio, stop_event=None):
    
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    
    obras = leer_obras_excel(ruta, stop_event=stop_event)
    if check_cancel():
            return {"cancelado": True}
    obras = consolidar_dh_por_ipn(obras)
    alertas = revisar_porcentajes(obras)
        
    cantidad_obras = len(obras)

    res = export_excel(obras, folio, stop_event=stop_event)

    if res.get("ok"):
        return {"ok": True, "alertas": alertas, "cantidad_obras": cantidad_obras, "ruta_destino": res.get("ruta_destino")}
