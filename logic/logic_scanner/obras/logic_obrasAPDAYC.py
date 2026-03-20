import pdfplumber
import re
import csv
import os
import unicodedata
from collections import defaultdict, OrderedDict
from utils.config_manager import obtener_carpeta_base
from openpyxl import load_workbook

EXTS_EXCEL = (".xlsx", ".xlsm", ".xltx", ".xltm")

# =========================
# Modo 1
# =========================

LINE_RE = re.compile(
    r'^\s*'
    r'(?:(?P<prefix>.*?)(?=\d{1,2}(?:,\d{3})+-\d{1,2}\s))?'
    r'(?P<ipi>\d{1,2}(?:,\d{3})+-\d{1,2})\s+'
    r'(?P<rol>[A-Z]{1,3})\s+'
    r'(?P<nombre>[A-ZÁÉÍÓÚÑ ]+?)\s+'
    r'(?P<p1>\d{1,3}\.\d{2})%(?P<soc1>[A-Z]+)\s*\((?P<n1>\d+)\)\s+'
    r'(?P<p2>\d{1,3}\.\d{2})%(?P<soc2>[A-Z]+)\s*\((?P<n2>\d+)\)\s*'
    r'$'
)

TITLE_RE = re.compile(
    r'^\s*'
    r'\d{1,3}(?:,\d{3})*\s+'
    r'(?P<iswc>T-\d{3}\.\d{3}\.\d{3}-\d)\s+'
    r'(?P<title>.+?)'
    r'\s+(?:OT\b.*)?$'
)

def parse_lines(text: str, stop_event=None):

    def check_cancel():
        return stop_event is not None and stop_event.is_set()

    rows = []
    current_title = None
    current_iswc = None

    for raw in text.splitlines():

        if check_cancel():
            return {"cancelado": True}

        line = raw.strip()
        if not line:
            continue

        tm = TITLE_RE.match(line)
        if tm:
            current_title = tm.group("title")
            current_iswc = tm.group("iswc")
            continue

        d = parse_dh_line(line)
        if not d or not current_iswc:
            continue

        d["nombre_obra"] = current_title
        d["iswc"] = current_iswc
        rows.append(d)

    return rows

def agrupar_por_iswc(rows, stop_event=None):

    def check_cancel():
        return stop_event is not None and stop_event.is_set()

    obras = defaultdict(lambda: {"titulo": None, "iswc": None, "dh": []})

    for r in rows:
        if check_cancel():
            return {"cancelado": True}

        iswc = r["iswc"]
        obras[iswc]["titulo"] = r["nombre_obra"]
        obras[iswc]["iswc"] = iswc
        obras[iswc]["dh"].append(r)

    return dict(obras)

def consolidar_dh_por_ipi(obras: dict, stop_event=None) -> dict:

    def check_cancel():
        return stop_event is not None and stop_event.is_set()

    for iswc, obra in obras.items():
        if check_cancel():
            return {"cancelado": True}

        acumulado = OrderedDict()

        for r in obra["dh"]:
            if check_cancel():
                return {"cancelado": True}

            key = (r["ipi"], r["rol"])

            if key not in acumulado:
                acumulado[key] = dict(r)
            else:
                acumulado[key]["p_share"] += r["p_share"]
                acumulado[key]["m_share"] += r["m_share"]

        obra["dh"] = list(acumulado.values())

    return obras

def parse_dh_line(line: str) -> dict | None:
    m = LINE_RE.match(line.strip())
    if not m:
        return None

    d = m.groupdict()

    p_share = float(d["p1"].replace(",", "."))
    m_share = float(d["p2"].replace(",", ".")) if d["p2"] else 0.0

    return {
        "ipi": d["ipi"],
        "rol": d["rol"],
        "nombre": d["nombre"],
        "p_share": p_share,
        "m_share": m_share,
        "p_soc": d.get("soc1"),
        "m_soc": d.get("soc2"),
    }

def parse_ipi(p):
    p = (p or "").replace(",", "").replace("-", "")
    return p.lstrip("0")

def parse_name(titulo: str) -> str:
    return re.sub(r"\s*\(.*?\)", "", titulo).strip()

def normalizar_nombre(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto or "")
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^A-Za-z\s]", "", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto

def revisar_porcentajes(obras: dict):
    alertas = []

    for key, obra in obras.items():
        titulo = obra.get("titulo") or "(sin título)"
        iswc = obra.get("iswc") or "(sin ISWC)"

        suma_p = sum(float(r.get("p_share", 0.0) or 0.0) for r in obra["dh"])
        suma_m = sum(float(r.get("m_share", 0.0) or 0.0) for r in obra["dh"])

        suma_p = round(suma_p, 2)
        suma_m = round(suma_m, 2)

        if suma_p != 100 or (suma_m != 100 and suma_m > 0):
            alertas.append((key, titulo, iswc, suma_p, suma_m))

    return alertas

# =========================
# Modo 2
# =========================

def extraer_excels(ruta):
    return [
        os.path.join(ruta, f)
        for f in os.listdir(ruta)
        if os.path.isfile(os.path.join(ruta, f)) and f.lower().endswith(EXTS_EXCEL)
    ]

def to_float(v):
    if v is None:
        return 0.0
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except:
        return 0.0

def norm_nombre_key(s: str) -> str:
    s = (s or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Za-z\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip().upper()
    return s

def leer_excel_por_headers(path_xlsx):
    wb = load_workbook(path_xlsx, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    req = ["COD_WORK_SQ", "TITLE", "ISWC", "IP_NAME", "NOMBRE", "COD_ROLE", "PORCENT_PER", "PORCENT_MEC"]
    faltan = [h for h in req if h not in idx]
    if faltan:
        raise ValueError(f"{os.path.basename(path_xlsx)}: faltan columnas {faltan}")

    rows = []

    last_work = ""
    last_title = ""
    last_iswc = ""

    for row in ws.iter_rows(min_row=2, values_only=True):
        work_raw = row[idx["COD_WORK_SQ"]]
        title_raw = row[idx["TITLE"]]
        iswc_raw = row[idx["ISWC"]]

        work = str(work_raw).strip() if work_raw is not None and str(work_raw).strip() != "" else last_work
        titulo = str(title_raw).strip() if title_raw is not None and str(title_raw).strip() != "" else last_title
        iswc = str(iswc_raw).strip() if iswc_raw is not None and str(iswc_raw).strip() != "" else last_iswc

        if work:
            last_work = work
        if titulo:
            last_title = titulo
        if iswc:
            last_iswc = iswc

        if not work:
            continue

        ipi = str(row[idx["IP_NAME"]] or "").replace("\xa0", "").strip()

        rows.append({
            "work_number": work,
            "titulo": titulo,
            "rol": str(row[idx["COD_ROLE"]] or "").strip(),
            "nombre": str(row[idx["NOMBRE"]] or "").strip(),
            "ipi": ipi,
            "p_share": to_float(row[idx["PORCENT_PER"]]),
            "m_share": to_float(row[idx["PORCENT_MEC"]]),
            "iswc": iswc,
        })

    return rows

def agrupar_por_worknumber(rows):
    obras = defaultdict(lambda: {"titulo": None, "work_number": None, "iswc": None, "dh": []})

    for r in rows:
        key = r["work_number"]
        obras[key]["titulo"] = r["titulo"]
        obras[key]["work_number"] = key
        obras[key]["iswc"] = r.get("iswc")
        obras[key]["dh"].append(r)

    return dict(obras)

def consolidar_dh_modo_2(obras):
    for _, obra in obras.items():
        acc = OrderedDict()

        for r in obra["dh"]:
            ipi = (r.get("ipi") or "").strip()

            if ipi and set(ipi) != {"0"}:
                key = ("IPI", ipi, r.get("rol"))
            else:
                key = ("NOMBRE", r.get("rol"), norm_nombre_key(r.get("nombre")))

            if key not in acc:
                acc[key] = dict(r)
                if not ipi or set(ipi) == {"0"}:
                    acc[key]["ipi"] = ""
            else:
                acc[key]["p_share"] += r.get("p_share", 0.0)
                acc[key]["m_share"] += r.get("m_share", 0.0)

        obra["dh"] = list(acc.values())

    return obras

# =========================
# Scanner
# =========================

def scanner(ruta, folio, stop_event=None, modo=1):

    def check_cancel():
        return stop_event is not None and stop_event.is_set()

    if modo == 1:
        with pdfplumber.open(ruta) as pdf:

            if check_cancel():
                return {"cancelado": True}

            texto = "\n".join(page.extract_text() or "" for page in pdf.pages)

            rows = parse_lines(texto, stop_event=stop_event)
            if isinstance(rows, dict) and rows.get("cancelado"):
                return rows

            obras = agrupar_por_iswc(rows, stop_event=stop_event)
            if isinstance(obras, dict) and obras.get("cancelado"):
                return obras

            obras = consolidar_dh_por_ipi(obras, stop_event=stop_event)
            if isinstance(obras, dict) and obras.get("cancelado"):
                return obras

            alertas = revisar_porcentajes(obras)
            cantidad_obras = len(obras)

    else:
        if check_cancel():
            return {"cancelado": True}

        rutas = extraer_excels(ruta)
        rows_global = []

        for ruta_excel in rutas:
            try:
                rows_global.extend(leer_excel_por_headers(ruta_excel))
            except Exception as e:
                print(f"⚠️ Saltando {ruta_excel}: {e}")

        obras = agrupar_por_worknumber(rows_global)
        obras = consolidar_dh_modo_2(obras)
        cantidad_obras = len(obras)
        alertas = revisar_porcentajes(obras)

    res = export_excel(obras, folio, stop_event=stop_event)

    if res.get("ok"):
        return {
            "ok": True,
            "alertas": alertas,
            "cantidad_obras": cantidad_obras,
            "ruta_destino": res.get("ruta_destino")
        }

# =========================
# Export
# =========================

def export_excel(obras, folio, stop_event=None):

    def check_cancel():
        return stop_event is not None and stop_event.is_set()

    carpeta_base = obtener_carpeta_base()
    ruta_destino = carpeta_base / "MMs" / f"mm_apdayc_{folio}.csv"

    fila_por_obra = {}
    fila = 1

    with open(
        ruta_destino,
        "w",
        newline="",
        encoding="latin-1"
    ) as f:
        writer = csv.writer(f, delimiter=";")

        for key, obra in obras.items():

            if check_cancel():
                return {"cancelado": True}

            fila_por_obra[key] = fila

            writer.writerow([parse_name(obra["titulo"]), "", "", "", "", ""])
            writer.writerow([])
            writer.writerow([])
            writer.writerow([])

            suma_m_obra = round(sum((x.get("m_share") or 0.0) for x in obra["dh"]), 2)
            mostrar_m = suma_m_obra > 0

            for r in obra["dh"]:
                m_cell = (
                    f"{r['m_share']:.2f}".replace(".", ",")
                    if mostrar_m
                    else ""
                )

                writer.writerow([
                    r["rol"],
                    normalizar_nombre(r["nombre"]),
                    f"{r['p_share']:.2f}".replace(".", ","),
                    m_cell,
                    "",
                    parse_ipi(r["ipi"]),
                ])

            writer.writerow([])

    return {"fila_por_obra": fila_por_obra, "ok": True, "ruta_destino": ruta_destino}