import os, re, csv, unicodedata
from openpyxl import load_workbook
from collections import defaultdict, OrderedDict
from utils.config_manager import obtener_carpeta_base

EXTS_EXCEL = (".xlsx", ".xlsm", ".xltx", ".xltm")

def extraer_excels(ruta):
    return [
        os.path.join(ruta, f)
        for f in os.listdir(ruta)
        if os.path.isfile(os.path.join(ruta, f)) and f.lower().endswith(EXTS_EXCEL)
    ]

def hoja_a_tabla(ws, max_rows=None, max_cols=None):
    rows = []
    for i, row in enumerate(ws.values, start=1):
        if max_rows and i > max_rows:
            break
        if max_cols:
            row = row[:max_cols]
        rows.append(list(row))
    while rows and all(v is None or str(v).strip() == "" for v in rows[-1]):
        rows.pop()
    return rows

def tabla_a_texto(tabla):
    lineas = []
    for row in tabla:
        celdas = [str(v).strip() for v in row if v is not None and str(v).strip() != ""]
        if celdas:
            lineas.append(" ".join(celdas))
    return "\n".join(lineas)

LINE_RE = re.compile(
    r'^(?P<nombre>.+?)\s+'
    r'(?P<rol>[A-Z]{1,3})\s+'
    r'(?P<ipn>\d{8,12})\s+'
    r'\d+\s+\d+\s+'
    r'(?P<p_share>\d+(?:[.,]\d+)?)\s+'
    r'\d+\s+'
    r'(?P<m_share>\d+(?:[.,]\d+)?)\s*$'
)

def extraer_titulo_afitap(text: str) -> str | None:
    lines = [ln.strip() for ln in text.splitlines()]
    for i, ln in enumerate(lines):
        if re.match(r'^---\s*HOJA:\s*Report\s*---\s*$', ln, re.IGNORECASE):
            for j in range(i + 1, min(i + 10, len(lines))):
                if lines[j]:
                    return lines[j]
    return None

def parse_lines(text: str):
    rows = []

    current_title = extraer_titulo_afitap(text)

    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue

        d = parse_dh_line(line)
        if not d or not current_title:
            continue

        d["nombre_obra"] = current_title
        rows.append(d)

    return rows

def agrupar_por_titulo(rows):
    obras = defaultdict(lambda: {"titulo": None, "dh": []})

    for r in rows:
        titulo = r["nombre_obra"]

        obras[titulo]["titulo"] = titulo
        obras[titulo]["dh"].append(r)

    return dict(obras)

def parse_dh_line(line: str) -> dict | None:
    m = LINE_RE.match(line.strip())
    if not m:
        return None

    d = m.groupdict()

    d["p_share"] = float(d["p_share"].replace(",", "."))

    if d["m_share"]:
        d["m_share"] = float(d["m_share"].replace(",", "."))
    else:
        d["m_share"] = 0.0
        d["m_soc"] = None  

    return d

def consolidar_dh_por_ipn(obras: dict) -> dict:
    
    for titulo, obra in obras.items():
        acumulado = OrderedDict()

        for r in obra["dh"]:
            key = (r["ipn"], r["rol"]) 

            if key not in acumulado:
                acumulado[key] = dict(r) 
            else:
                acumulado[key]["p_share"] += r["p_share"]
                acumulado[key]["m_share"] += r["m_share"]

        obra["dh"] = list(acumulado.values())

    return obras

def revisar_porcentajes(obras: dict):
    alertas = []

    for key, obra in obras.items():
        titulo = obra.get("titulo") or "(sin título)"
        iswc = obra.get("iswc") or "(sin ISWC)"

        suma_p = sum(r.get("p_share", 0.0) for r in obra["dh"])
        suma_m = sum(r.get("m_share", 0.0) for r in obra["dh"])

        suma_p = round(suma_p, 2)
        suma_m = round(suma_m, 2)
        
        if suma_p != 100 or (suma_m != 100 and suma_m > 0):
            alertas.append((key, titulo, iswc, suma_p, suma_m))

    return alertas

def scanner_excel(carpeta_excels, folio, stop_event=None):
    
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    
    rutas = extraer_excels(carpeta_excels)

    all_rows_global = []

    for ruta in rutas:
        if check_cancel():
            return {"cancelado": True}
        try:
            wb = load_workbook(ruta, data_only=True)
        except Exception as e:
            print(f"⚠️ No pude abrir: {ruta}\n   Error: {e}")
            continue

        for nombre_hoja in wb.sheetnames:
            if check_cancel():
                return {"cancelado": True}
            ws = wb[nombre_hoja]
            tabla = hoja_a_tabla(ws)
            if not tabla:
                continue

            texto = f"--- HOJA: {nombre_hoja} ---\n" + tabla_a_texto(tabla)
            rows = parse_lines(texto)
            all_rows_global.extend(rows)

    obras = agrupar_por_titulo(all_rows_global)
    obras = consolidar_dh_por_ipn(obras)
    alertas = revisar_porcentajes(obras)
    cantidad_obras = len(obras)
    
    res = export_excel(obras, folio, stop_event=stop_event)
    
    if res.get("ok"):
        return {"ok": True, "alertas": alertas, "cantidad_obras": cantidad_obras, "ruta_destino": res.get("ruta_destino")}

def parse_ipi(p):
    return p.lstrip("0")

def parse_name(titulo: str) -> str:
    return re.sub(r"\s*\(.*?\)", "", titulo).strip()

def normalizar_nombre(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto)
    
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    
    texto = re.sub(r"[^A-Za-z\s]", "", texto)
    
    texto = re.sub(r"\s+", " ", texto).strip()

    return texto

def export_excel(obras, folio, stop_event=None):
    
    def check_cancel():
        return stop_event is not None and stop_event.is_set()
    
    carpeta_base = obtener_carpeta_base()
    
    ruta_destino = carpeta_base / "MMs" / f"mm_mesam_{folio}.csv"
    
    fila_por_obra = {}
    fila = 1

    with open(
        ruta_destino,
        "w",
        newline="",
        encoding="latin-1"
    ) as f:
        writer = csv.writer(f, delimiter=";")

        for key, obra in obras.items():
            
            if check_cancel():
                return {"cancelado": True}
            
            fila_por_obra[key] = fila
            
            # Fila título
            writer.writerow([parse_name(obra["titulo"]), "", "", "", "", ""])
            writer.writerow([])
            writer.writerow([])
            writer.writerow([])
            
            suma_m_obra = round(sum((x.get("m_share") or 0.0) for x in obra["dh"]), 2)
            mostrar_m = suma_m_obra > 0

            # DH
            for r in obra["dh"]:
                
                if check_cancel():
                    return {"cancelado": True}
                
                m_cell = (
                    f"{r['m_share']:.2f}".replace(".", ",")
                    if mostrar_m
                    else ""
                )
                
                writer.writerow([
                    r["rol"],
                    normalizar_nombre(r["nombre"]),
                    f"{r['p_share']:.2f}".replace(".", ","),
                    m_cell,
                    "",
                    parse_ipi(r["ipn"]),
                ])

            writer.writerow([])
            
    return {"fila_por_obra": fila_por_obra, "ok": True, "ruta_destino": ruta_destino}


