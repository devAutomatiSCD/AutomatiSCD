from openpyxl import load_workbook
from pathlib import Path
import re, unicodedata, csv
from collections import defaultdict

# Si quieres seguir usando tu carpeta base:
# from utils.config_manager import obtener_carpeta_base


def normalize_title(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("&", " y ").replace("Ø", "o").replace("ø", "o")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^0-9A-Za-z ]+", " ", s)   # fuera signos raros
    s = re.sub(r"\s{2,}", " ", s).strip().upper()
    return s


def normalize_name(s: str) -> str:
    if s is None:
        return ""
    s = str(s).replace("&", " y ").replace("Ø", "o").replace("ø", "o")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s{2,}", " ", s).strip().upper()
    return s


def clean_ipi(val) -> str:
    """Quita ceros iniciales (y limpia basura)."""
    if val is None:
        return ""
    s = str(val).strip()
    # por si aparecen prefijos tipo T- o D-
    if re.match(r"^[Tt]-", s):
        s = s[2:]
    if re.match(r"^[Dd]-", s):
        s = s[2:]
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^\d]", "", s)
    s = s.lstrip("0")
    return s or "0"


def parse_pct_to_bp(val):
    """Convierte % a centésimas (basis points). 100% => 10000."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(round(float(val) * 100))

    s = str(val).strip()
    if not s:
        return None
    s = s.replace("%", "").replace(" ", "").replace(",", ".")
    try:
        return int(round(float(s) * 100))
    except:
        return None


def bp_to_str(bp: int) -> str:
    """Centésimas -> string con coma decimal, sin ceros inútiles."""
    v = bp / 100
    s = f"{v:.2f}".rstrip("0").rstrip(".")
    return s.replace(".", ",")


def is_029(x) -> bool:
    if x is None:
        return False
    s = str(x).strip()
    if s == "":
        return False
    digits = re.sub(r"\D", "", s)
    if digits == "":
        return False
    return digits.zfill(3) == "029"


def repartir_bp_igual(total_bp: int, rows: list[dict]):
    """Reparte total_bp en partes iguales (con resto distribuido 1 a 1)."""
    n = len(rows)
    if n <= 0 or total_bp == 0:
        return
    q, rem = divmod(total_bp, n)
    for i, r in enumerate(rows):
        r["pct_bp"] = (r["pct_bp"] or 0) + q + (1 if i < rem else 0)


def ajustar_a_100(dh_rows: list[dict]):
    """
    Deja la obra en 100,00% exacto:
      - si falta poco: suma al menor
      - si sobra poco: descuenta al mayor
      - si falta MUCHO (obra <= 95%): reparte entre todos
    """
    valid = [r for r in dh_rows if r.get("pct_bp") is not None]
    if not valid:
        return

    total = sum(r["pct_bp"] for r in valid)
    target = 10000
    diff = target - total  # >0 falta, <0 sobra

    if diff == 0:
        return

    if diff > 0:
        # falta %
        if total <= 9500:   # 95% o menos: repartir entre todos
            repartir_bp_igual(diff, valid)
        else:
            # falta poco: dárselo al que tiene menos
            rmin = min(valid, key=lambda r: r["pct_bp"])
            rmin["pct_bp"] += diff
    else:
        # sobra %: quitárselo al/los que más tienen
        sobra = -diff
        for r in sorted(valid, key=lambda r: r["pct_bp"], reverse=True):
            if sobra <= 0:
                break
            take = min(r["pct_bp"], sobra)
            r["pct_bp"] -= take
            sobra -= take


def escribir_csv_estructurado(records, out_csv: Path, delimiter=";"):
    """
    Estructura:
      OBRA
      (3 filas vacías)
      ROLE;NOMBRE;%;IPI
      ...
      (1 fila vacía)
    """
    obras = defaultdict(list)
    for r in records:
        obras[r["OBRA"]].append(r)

    with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=delimiter)
        for obra, filas in obras.items():
            w.writerow([obra])
            w.writerow([]); w.writerow([]); w.writerow([])

            for r in filas:
                w.writerow([
                    r["ROLECODE"],     # A
                    r["NOMBRE_DH"],    # B
                    r["PORCENTAJE"],   # C
                    "",                # D (Soc Adm eliminado)
                    "",                # E vacío
                    r["IPI"]           # F
                ])

            w.writerow([])


def procesar_excel_a_csv(ruta_excel: str):
    ruta_excel = Path(ruta_excel)
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    def row_vals(r):
        return [ws.cell(r, c).value for c in range(1, 7)]  # A..F

    records = []
    current_work = None
    dh_rows = []

    def flush_work(work, dhs):
        if not work or not dhs:
            return

        # Regla 029: eliminar DH con 029 y repartir su % en partes iguales entre editores (rol contiene "E")
        rows_029 = [r for r in dhs if is_029(r.get("soc_adm"))]
        if rows_029:
            removed_bp = sum((r.get("pct_bp") or 0) for r in rows_029)
            dhs = [r for r in dhs if not is_029(r.get("soc_adm"))]

            editors = [r for r in dhs if "E" in str(r.get("role", "")).upper()]
            repartir_bp_igual(removed_bp, editors)

        # Ajuste exacto a 100%
        ajustar_a_100(dhs)

        obra_norm = normalize_title(work)

        for r in dhs:
            pct = "" if r.get("pct_bp") is None else bp_to_str(r["pct_bp"])
            records.append({
                "OBRA": obra_norm,
                "ROLECODE": str(r.get("role", "")).strip().upper(),
                "NOMBRE_DH": normalize_name(r.get("name", "")),
                "PORCENTAJE": pct,
                "IPI": r.get("ipi_clean", ""),  # <- IPI como texto, NO porcentaje
            })

        return dhs  # por si quieres debug

    # Saltamos fila 1 (encabezados)
    for r in range(2, ws.max_row + 1):
        a, b, c, d, e, f = row_vals(r)

        is_blank = all(v is None or (isinstance(v, str) and v.strip() == "") for v in [a, b, c, d, e, f])
        if is_blank:
            continue

        # Fila de obra: solo A tiene valor
        if a is not None and all(x is None or (isinstance(x, str) and x.strip() == "") for x in [b, c, d, e, f]):
            flush_work(current_work, dh_rows)
            current_work = str(a)
            dh_rows = []
            continue
        
        missing_ipi_flags = [] 

        # Fila DH: A=rol, B=nombre, C=%, D=soc_adm, F=ipi
        if a is not None and b is not None:
            role = str(a).strip()
            name = b
            pct_bp = parse_pct_to_bp(c)
            soc_adm = d
            ipi_clean = clean_ipi(f)

            # NUEVO: detectar IPI faltante
            if ipi_clean in ("", "0"):
                missing_ipi_flags.append({
                    "row": r,
                    "obra": current_work,
                    "name": str(name).strip() if name is not None else "",
                    "role": role
                })

            dh_rows.append({
                "role": role,
                "name": name,
                "pct_bp": pct_bp,
                "soc_adm": soc_adm,
                "ipi_clean": ipi_clean,
                "excel_row": r,     # <- NUEVO: guardo fila original por si la quieres después
            })

    flush_work(current_work, dh_rows)

    # Salida: mismo directorio que el excel (cámbialo a tu carpeta_base si quieres)
    out_csv = ruta_excel.with_name(f"{ruta_excel.stem}_limpio.csv")

    # Si quieres tu ruta base:
    # carpeta_base = obtener_carpeta_base()
    # out_csv = carpeta_base / "MMs procesados" / f"{ruta_excel.stem}_limpio.csv"

    escribir_csv_estructurado(records, out_csv, delimiter=";")
    print(missing_ipi_flags)
    return {"ok": True, "ruta_salida":out_csv, "alerta": missing_ipi_flags}